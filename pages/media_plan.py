"""Media Plan page."""
import calendar
import io
import json
import os
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as _components
import toml

# ── Constants ─────────────────────────────────────────────────────────────────

MARKET_LABELS = {
    'AT': 'Austria',
    'BE': 'Belgium',
    'BG': 'Bulgaria',
    'HR': 'Croatia',
    'CY': 'Cyprus',
    'CZ': 'Czech Republic',
    'DK': 'Denmark',
    'EE': 'Estonia',
    'FI': 'Finland',
    'FR': 'France',
    'DE': 'Germany',
    'GR': 'Greece',
    'HU': 'Hungary',
    'IE': 'Ireland',
    'IT': 'Italy',
    'LV': 'Latvia',
    'LT': 'Lithuania',
    'LU': 'Luxembourg',
    'MT': 'Malta',
    'NL': 'Netherlands',
    'NO': 'Norway',
    'PL': 'Poland',
    'PT': 'Portugal',
    'RO': 'Romania',
    'SK': 'Slovakia',
    'SI': 'Slovenia',
    'ES': 'Spain',
    'SE': 'Sweden',
    'CH': 'Switzerland',
    'UK': 'United Kingdom',
}

# Default benchmark values — markets with known data have specific values,
# others fall back to regional averages (Western EU vs Eastern EU).
# CPM YouTube | CPM LinkedIn | View Rate | CTR YT | CTR LI | Frequency
# CPC Search  | CTR Search   | Click→Session | Conv Rate

def _default_bench(cpm_yt, cpm_li, view_rate, ctr_yt, ctr_li, freq,
                   cpc_s, ctr_s, c2s_yt=0.80, c2s_li=0.82, c2s_s=0.86,
                   cr=0.02, l2m=0.20, m2s=0.30):
    shared = {'conv_rate': cr, 'lead_to_mql': l2m, 'mql_to_sql': m2s}
    return {
        'YouTube':  {'cpm': cpm_yt, 'view_rate': view_rate, 'ctr': ctr_yt,
                     'frequency': freq, 'click_to_session': c2s_yt, **shared},
        'LinkedIn': {'cpm': cpm_li, 'ctr': ctr_li,
                     'frequency': freq, 'click_to_session': c2s_li, **shared},
        'Search':   {'cpc': cpc_s,  'ctr': ctr_s,
                     'click_to_session': c2s_s, **shared},
    }

BENCH = {
    'AT': _default_bench(11.0, 18.0, 0.31, 0.0035, 0.0038, 3.0, 2.30, 0.030),
    'BE': _default_bench(11.0, 17.0, 0.31, 0.0035, 0.0038, 3.0, 2.20, 0.030),
    'BG': _default_bench(3.5,  7.0,  0.30, 0.0028, 0.0032, 3.0, 0.70, 0.022),
    'HR': _default_bench(4.5,  9.0,  0.30, 0.0030, 0.0033, 3.0, 0.90, 0.023),
    'CY': _default_bench(7.0, 13.0,  0.30, 0.0030, 0.0035, 3.0, 1.50, 0.026),
    'CZ': _default_bench(5.0,  9.5,  0.30, 0.0030, 0.0033, 3.0, 1.00, 0.024),
    'DK': _default_bench(13.0, 21.0, 0.31, 0.0035, 0.0040, 3.0, 2.80, 0.032),
    'EE': _default_bench(4.5,  9.0,  0.30, 0.0030, 0.0033, 3.0, 0.90, 0.023),
    'FI': _default_bench(12.0, 19.0, 0.31, 0.0035, 0.0038, 3.0, 2.50, 0.030),
    'FR': _default_bench(10.0, 17.0, 0.31, 0.0033, 0.0038, 3.0, 2.00, 0.028),
    'DE': _default_bench(11.0, 18.0, 0.31, 0.0035, 0.0038, 3.0, 2.20, 0.030),
    'GR': _default_bench(5.0,  10.0, 0.30, 0.0028, 0.0032, 3.0, 1.00, 0.024),
    'HU': _default_bench(4.0,  8.0,  0.30, 0.0028, 0.0032, 3.0, 0.80, 0.022),
    'IE': _default_bench(12.0, 19.0, 0.31, 0.0035, 0.0040, 3.0, 2.50, 0.032),
    'IT': _default_bench(8.0,  14.0, 0.30, 0.0030, 0.0035, 3.0, 1.60, 0.026),
    'LV': _default_bench(4.5,  9.0,  0.30, 0.0030, 0.0033, 3.0, 0.90, 0.023),
    'LT': _default_bench(4.5,  9.0,  0.30, 0.0030, 0.0033, 3.0, 0.90, 0.023),
    'LU': _default_bench(12.0, 18.0, 0.31, 0.0035, 0.0038, 3.0, 2.40, 0.030),
    'MT': _default_bench(7.0,  13.0, 0.30, 0.0030, 0.0035, 3.0, 1.40, 0.025),
    'NL': _default_bench(10.0, 16.0, 0.31, 0.0035, 0.0040, 3.0, 2.00, 0.030),
    'NO': _default_bench(13.0, 22.0, 0.31, 0.0035, 0.0040, 3.0, 2.80, 0.032),
    'PL': _default_bench( 5.0,  9.0, 0.30, 0.0030, 0.0035, 3.0, 1.00, 0.025),
    'PT': _default_bench( 6.0, 11.0, 0.30, 0.0030, 0.0035, 3.0, 1.20, 0.025),
    'RO': _default_bench( 3.5,  7.0, 0.30, 0.0028, 0.0032, 3.0, 0.70, 0.022),
    'SK': _default_bench( 4.5,  9.0, 0.30, 0.0030, 0.0033, 3.0, 0.90, 0.023),
    'SI': _default_bench( 5.0, 10.0, 0.30, 0.0030, 0.0033, 3.0, 1.00, 0.024),
    'ES': _default_bench( 6.0, 12.0, 0.30, 0.0030, 0.0035, 3.0, 1.50, 0.025),
    'SE': _default_bench(12.0, 20.0, 0.31, 0.0035, 0.0040, 3.0, 2.60, 0.031),
    'CH': _default_bench(13.0, 21.0, 0.31, 0.0035, 0.0040, 3.0, 2.90, 0.032),
    'UK': _default_bench(12.0, 20.0, 0.31, 0.0035, 0.0040, 3.0, 2.50, 0.035),
}

CH_COLORS = {
    'YouTube':  ['#1F497D', '#437CA3', '#6B9FBF', '#9FC3D5', '#C5DCE8'],
    'LinkedIn': ['#1F6152', '#2E8A72', '#4DB896', '#7DCFB0', '#A8E4D0'],
    'Search':   ['#4285F4', '#5A95F5', '#74A5F6', '#8EB5F7', '#A8C5F8'],
}

ALL_GOALS = ['Awareness', 'Traffic', 'Conversion']

MARKET_GROUPS = {
    'DACH':        ['DE', 'AT', 'CH'],
    'Nordics':     ['DK', 'FI', 'NO', 'SE'],
    'BeNeLux':     ['BE', 'NL', 'LU'],
    'Southern EU': ['ES', 'IT', 'PT', 'GR'],
    'CEE':         ['PL', 'CZ', 'HU', 'RO', 'BG', 'SK', 'SI', 'HR'],
    'UK + IE':     ['UK', 'IE'],
    'All EU':      ['DE','AT','CH','DK','FI','NO','SE','BE','NL','LU','ES','IT','PT','GR',
                    'PL','CZ','HU','RO','BG','SK','SI','HR','FR','UK','IE','EE','LV','LT',
                    'CY','MT'],
}

PLAN_TEMPLATES = {
    'DACH Awareness Launch': {
        'markets': ['DE','AT','CH'], 'budget': 50000,
        'goals': {'Awareness': ['YouTube']},
    },
    'Pan-EU Lead Gen': {
        'markets': ['DE','FR','NL','BE','ES','IT','SE','PL'], 'budget': 120000,
        'goals': {'Traffic': ['LinkedIn','Search'], 'Conversion': ['Search']},
    },
    'Single Market Full Funnel': {
        'markets': ['DE'], 'budget': 80000,
        'goals': {'Awareness': ['YouTube'], 'Traffic': ['LinkedIn','Search'], 'Conversion': ['Search']},
    },
    'Nordics Brand Building': {
        'markets': ['DK','FI','NO','SE'], 'budget': 60000,
        'goals': {'Awareness': ['YouTube','LinkedIn']},
    },
    'BeNeLux Performance': {
        'markets': ['BE','NL','LU'], 'budget': 40000,
        'goals': {'Traffic': ['Search','LinkedIn'], 'Conversion': ['Search']},
    },
}

BENCH_PRESET_FACTORS = {
    'Conservative': {'cpm': 1.15, 'cpc': 1.15, 'ctr': 0.80, 'view_rate': 0.85, 'conv_rate': 0.75},
    'Average':      {'cpm': 1.00, 'cpc': 1.00, 'ctr': 1.00, 'view_rate': 1.00, 'conv_rate': 1.00},
    'Aggressive':   {'cpm': 0.88, 'cpc': 0.88, 'ctr': 1.25, 'view_rate': 1.15, 'conv_rate': 1.30},
}

BENCH_HELP = {
    'cpm':              'Cost per 1,000 impressions. Western EU: €10–13. Eastern EU: €3.5–5.',
    'cpc':              'Cost per click on Search. Ranges €0.70 (Eastern EU) to €2.90 (DACH/Nordics).',
    'ctr':              'Click-through rate — % of ad impressions that result in a click.',
    'view_rate':        'YouTube: % of impressions resulting in a 30-second (or full-video) view.',
    'frequency':        'Average number of times one unique user sees your ad across the campaign.',
    'click_to_session': '% of clicks that result in a tracked site session (accounts for pixel gaps and bounces).',
    'conv_rate':        '% of sessions that convert to a lead. Typical B2B range: 1–5%.',
    'lead_to_mql':      '% of raw leads that qualify as Marketing Qualified Leads. Typical B2B: 15–35%.',
    'mql_to_sql':       '% of MQLs accepted by sales as Sales Qualified Leads. Typical B2B: 20–40%.',
}

DONUT_PALETTE = ['#2BB5A5','#1F497D','#4DB896','#437CA3','#5A3E7A','#E8A838','#8B3A3A','#6B7280',
                 '#229990','#2E8A72','#6B9FBF','#7DCFB0']

ADDITIVE = ['Budget', 'impressions', 'reach', 'views', 'clicks', 'sessions',
            'conversions', 'mql', 'sql']

COL_FMT = {
    'Budget':           ('Spent (€)',        lambda x: f'€{x:,.0f}'),
    'impressions':      ('Impressions',      lambda x: f'{int(round(x)):,}'),
    'eff_cpm':          ('CPM (€)',          lambda x: f'€{x:.2f}'),
    'reach':            ('Reach',            lambda x: f'{int(round(x)):,}'),
    'views':            ('Views',            lambda x: f'{int(round(x)):,}'),
    'clicks':           ('Clicks',           lambda x: f'{int(round(x)):,}'),
    'cpc':              ('CPC (€)',          lambda x: f'€{x:.2f}'),
    'ctr':              ('CTR',              lambda x: f'{x*100:.2f}%'),
    'click_to_session': ('Click→Session %',  lambda x: f'{x*100:.0f}%'),
    'sessions':         ('Sessions',         lambda x: f'{int(round(x)):,}'),
    'conv_rate':        ('Session→Lead %',   lambda x: f'{x*100:.2f}%'),
    'conversions':      ('Leads',            lambda x: f'{int(round(x)):,}'),
    'cpa':              ('Cost per Lead (€)', lambda x: f'€{x:,.2f}'),
    'lead_to_mql':      ('Lead→MQL %',       lambda x: f'{x*100:.0f}%'),
    'mql':              ('MQL',              lambda x: f'{int(round(x)):,}'),
    'cost_per_mql':     ('CPMQL (€)',        lambda x: f'€{x:,.2f}'),
    'mql_to_sql':       ('MQL→SQL %',        lambda x: f'{x*100:.0f}%'),
    'sql':              ('SQL',              lambda x: f'{int(round(x)):,}'),
    'cost_per_sql':     ('CPSQL (€)',        lambda x: f'€{x:,.2f}'),
    'view_rate':        ('View Rate',        lambda x: f'{x*100:.1f}%'),
    'cpv':              ('CPV (€)',          lambda x: f'€{x:.2f}'),
    'cvr':              ('Click→Lead %',     lambda x: f'{x*100:.2f}%'),
    'click_to_session_raw': ('Click→Session', lambda x: f'{x*100:.0f}%'),
}

# Columns shown per (channel, goal), in the exact order the user specified.
_TRAFFIC_COLS    = ['Budget', 'impressions', 'eff_cpm', 'clicks', 'cpc',
                    'ctr', 'click_to_session', 'sessions']
_CONVERSION_COLS = ['Budget', 'impressions', 'eff_cpm', 'clicks', 'cpc',
                    'ctr', 'click_to_session', 'sessions',
                    'conv_rate', 'conversions', 'cpa',
                    'lead_to_mql', 'mql', 'cost_per_mql',
                    'mql_to_sql', 'sql', 'cost_per_sql']

PHASE_COLS = {
    ('YouTube', 'Awareness'):   ['Budget', 'impressions', 'reach', 'views', 'cpv', 'ctr', 'clicks', 'cpc'],
    ('YouTube', 'Traffic'):     _TRAFFIC_COLS,
    ('YouTube', 'Conversion'):  _CONVERSION_COLS,
    ('Search',  'Awareness'):   ['Budget', 'impressions', 'ctr', 'clicks', 'cpc'],
    ('Search',  'Traffic'):     _TRAFFIC_COLS,
    ('Search',  'Conversion'):  _CONVERSION_COLS,
    ('LinkedIn','Awareness'):   ['Budget', 'impressions', 'reach', 'ctr', 'clicks', 'cpc'],
    ('LinkedIn','Traffic'):     _TRAFFIC_COLS,
    ('LinkedIn','Conversion'):  _CONVERSION_COLS,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def generate_periods(start, end, breakdown):
    periods, cur = [], start
    if breakdown == 'Daily':
        while cur <= end:
            periods.append({'label': cur.strftime('%b %d, %Y'), 'days': 1})
            cur += timedelta(days=1)
    elif breakdown == 'Weekly':
        while cur <= end:
            p_end = min(cur + timedelta(days=6), end)
            periods.append({'label': f"{cur.strftime('%b %d')} – {p_end.strftime('%b %d')}", 'days': (p_end - cur).days + 1})
            cur += timedelta(days=7)
    elif breakdown == 'Bi-Weekly':
        while cur <= end:
            p_end = min(cur + timedelta(days=13), end)
            periods.append({'label': f"{cur.strftime('%b %d')} – {p_end.strftime('%b %d')}", 'days': (p_end - cur).days + 1})
            cur += timedelta(days=14)
    elif breakdown == 'Monthly':
        while cur <= end:
            last = calendar.monthrange(cur.year, cur.month)[1]
            p_end = min(date(cur.year, cur.month, last), end)
            periods.append({'label': cur.strftime('%B %Y'), 'days': (p_end - cur).days + 1})
            cur = date(cur.year + 1, 1, 1) if cur.month == 12 else date(cur.year, cur.month + 1, 1)
    return periods


def calc_row(budget, bm, goal, channel, conv_rate):
    if budget <= 0:
        return {'Budget': budget}
    r = {'Budget': budget}

    if channel == 'Search':
        cpc = bm.get('cpc', 2.0)
        ctr = bm.get('ctr', 0.03)
        c2s = bm.get('click_to_session', 0.85)
        if cpc <= 0:
            return r
        clicks = budget / cpc
        impressions = clicks / ctr if ctr > 0 else 0
        r.update({'impressions': impressions, 'clicks': clicks, 'cpc': cpc, 'ctr': ctr})
        if goal in ('Traffic', 'Conversion'):
            r['sessions'] = clicks * c2s
            r['click_to_session'] = c2s
        if goal == 'Conversion':
            convs = r['sessions'] * conv_rate
            r['conversions'] = convs
            r['cpa']       = budget / convs if convs > 0 else 0
            r['cvr']       = convs / clicks if clicks > 0 else 0
            r['conv_rate'] = conv_rate
            l2m = bm.get('lead_to_mql', 0.20)
            m2s = bm.get('mql_to_sql',  0.30)
            mql = convs * l2m
            sql = mql * m2s
            r['lead_to_mql']  = l2m
            r['mql']          = mql
            r['cost_per_mql'] = budget / mql if mql > 0 else 0
            r['mql_to_sql']   = m2s
            r['sql']          = sql
            r['cost_per_sql'] = budget / sql if sql > 0 else 0

    else:
        cpm = bm.get('cpm', 10.0)
        ctr = bm.get('ctr', 0.003)
        freq = bm.get('frequency', 3.0)
        c2s = bm.get('click_to_session', 0.80)
        if cpm <= 0:
            return r
        imp = (budget / cpm) * 1000
        reach = imp / freq if freq > 0 else 0
        clicks = imp * ctr
        r.update({
            'impressions': imp,
            'reach':       reach,
            'clicks':      clicks,
            'ctr':         ctr,
            'cpc':         budget / clicks if clicks > 0 else 0,
        })
        if goal == 'Awareness' and channel == 'YouTube':
            vr = bm.get('view_rate', 0.31)
            views = imp * vr
            r['views'] = views
            r['cpv']   = budget / views if views > 0 else 0
        if goal in ('Traffic', 'Conversion'):
            r['sessions']         = clicks * c2s
            r['click_to_session'] = c2s
        if goal == 'Conversion':
            convs = r['sessions'] * conv_rate
            r['conversions'] = convs
            r['cpa']         = budget / convs if convs > 0 else 0
            r['cvr']         = convs / clicks if clicks > 0 else 0
            r['conv_rate']   = conv_rate
            l2m = bm.get('lead_to_mql', 0.20)
            m2s = bm.get('mql_to_sql',  0.30)
            mql = convs * l2m
            sql = mql * m2s
            r['lead_to_mql']  = l2m
            r['mql']          = mql
            r['cost_per_mql'] = budget / mql if mql > 0 else 0
            r['mql_to_sql']   = m2s
            r['sql']          = sql
            r['cost_per_sql'] = budget / sql if sql > 0 else 0

    # Effective CPM — works for all channels once impressions are known
    if r.get('impressions', 0) > 0:
        r['eff_cpm'] = r['Budget'] / r['impressions'] * 1000

    return r


def build_table(periods, total_budget, bm, goal, channel, conv_rate):
    total_days = sum(p['days'] for p in periods) or 1
    rows = []
    for p in periods:
        bud = total_budget * p['days'] / total_days
        m = calc_row(bud, bm, goal, channel, conv_rate)
        rows.append({'Period': p['label'], 'Days': p['days'], **m})
    df = pd.DataFrame(rows)
    total_m = calc_row(total_budget, bm, goal, channel, conv_rate)
    total_df = pd.DataFrame([{'Period': 'TOTAL', 'Days': total_days, **total_m}])
    return pd.concat([df, total_df], ignore_index=True)


def fmt_df(df, ch=None, goal=None):
    out = df[['Period', 'Days']].copy()
    col_order = PHASE_COLS.get((ch, goal)) if ch and goal else None
    keys = col_order if col_order else list(COL_FMT.keys())
    for raw in keys:
        if raw in COL_FMT and raw in df.columns:
            label, fn = COL_FMT[raw]
            out[label] = df[raw].apply(fn)
    return out


def make_funnel(df, goal, channel, title):
    t = df[df['Period'] == 'TOTAL'].iloc[0]
    colors = CH_COLORS.get(channel, CH_COLORS['YouTube'])

    if channel == 'Search':
        if goal == 'Awareness':
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks')]
        elif goal == 'Traffic':
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions')]
        else:
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions'), ('Conversions', 'conversions')]
    elif channel == 'YouTube':
        if goal == 'Awareness':
            stages = [('Impressions', 'impressions'), ('Reach', 'reach'), ('Views', 'views'), ('Clicks', 'clicks')]
        elif goal == 'Traffic':
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions')]
        else:
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions'), ('Conversions', 'conversions')]
    else:  # LinkedIn
        if goal == 'Awareness':
            stages = [('Impressions', 'impressions'), ('Reach', 'reach'), ('Clicks', 'clicks')]
        elif goal == 'Traffic':
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions')]
        else:
            stages = [('Impressions', 'impressions'), ('Clicks', 'clicks'), ('Sessions', 'sessions'), ('Conversions', 'conversions')]

    labels = [s[0] for s in stages]
    values = [t.get(s[1], 0) for s in stages]

    fig = go.Figure(go.Funnel(
        y=labels,
        x=values,
        textinfo='value+percent initial',
        marker={'color': colors[:len(stages)]},
        textfont={'size': 10},
        connector={'line': {'color': 'rgba(0,0,0,0.1)', 'width': 1}},
    ))
    fig.update_layout(
        title={'text': title, 'font': {'size': 12, 'color': '#1F497D'}, 'x': 0.5, 'xanchor': 'center'},
        margin={'t': 40, 'b': 10, 'l': 10, 'r': 10},
        height=265,
        paper_bgcolor='rgba(0,0,0,0)',
    )
    return fig


def _duplicate_scenario(sid):
    import re as _re
    new_sid = len(st.session_state.scenario_names)
    st.session_state.scenario_names.append(st.session_state.scenario_names[sid] + ' (copy)')
    skip_prefixes = ('remove_', 'dup_', 'rename_', 'btn_', 'dl_', 'funnel_', 'grand_',
                     'preset_', 'eq_', 'cpm_eff_', 'grp_', 'tpl_', 'pacing_', 'bar_')
    pattern = _re.compile(rf'^(.+)_{sid}$')
    for k, v in list(st.session_state.items()):
        m = pattern.match(str(k))
        if m and not any(str(k).startswith(p) for p in skip_prefixes):
            st.session_state[f'{m.group(1)}_{new_sid}'] = v
    st.rerun()


def _apply_template_data(sid, tpl):
    """Apply a template dict to a scenario slot."""
    n = max(len(tpl['markets']), 1)
    st.session_state[f'selected_markets_{sid}'] = tpl['markets']
    st.session_state[f'total_budget_{sid}']     = tpl['budget']
    default_pct = round(100.0 / n, 1)
    for mkt in tpl['markets']:
        st.session_state[f'pct_{mkt}_{sid}'] = default_pct
    for goal in ALL_GOALS:
        goal_on = goal in tpl['goals']
        st.session_state[f'sb_goal_{goal}_{sid}'] = goal_on
        for ch, key in [('YouTube', 'sb_yt'), ('Search', 'sb_s'), ('LinkedIn', 'sb_li')]:
            st.session_state[f'{key}_{goal}_{sid}'] = goal_on and ch in tpl['goals'].get(goal, [])
    st.rerun()


def _apply_template(sid, tpl_name):
    _apply_template_data(sid, PLAN_TEMPLATES[tpl_name])


def _current_as_template(sid):
    """Snapshot the current scenario config as a saveable template dict."""
    ss = st.session_state
    goals = {}
    for goal in ALL_GOALS:
        if ss.get(f'sb_goal_{goal}_{sid}', False):
            chs = []
            if ss.get(f'sb_yt_{goal}_{sid}', False): chs.append('YouTube')
            if ss.get(f'sb_s_{goal}_{sid}',  False): chs.append('Search')
            if ss.get(f'sb_li_{goal}_{sid}', False): chs.append('LinkedIn')
            if chs:
                goals[goal] = chs
    return {
        'markets': ss.get(f'selected_markets_{sid}', []),
        'budget':  ss.get(f'total_budget_{sid}', 0),
        'goals':   goals,
    }


def _apply_bench_preset(ch, mkt, goal, sid, preset_name):
    b = BENCH[mkt][ch]
    f = BENCH_PRESET_FACTORS[preset_name]
    keys = []
    if ch == 'Search':
        keys = [
            (f'cpc_{mkt}_{ch}_{goal}_{sid}',              round(b['cpc'] * f['cpc'], 2)),
            (f'ctr_{mkt}_{ch}_{goal}_{sid}',              round(b['ctr'] * 100 * f['ctr'], 2)),
            (f'click_to_session_{mkt}_{ch}_{goal}_{sid}', round(b.get('click_to_session', 0.85) * 100, 0)),
            (f'conv_rate_{mkt}_{ch}_{goal}_{sid}',        round(b.get('conv_rate', 0.03) * 100 * f['conv_rate'], 1)),
        ]
    elif ch == 'YouTube':
        keys = [
            (f'cpm_{mkt}_{ch}_{goal}_{sid}',              round(b['cpm'] * f['cpm'], 2)),
            (f'view_rate_{mkt}_{ch}_{goal}_{sid}',        round(b.get('view_rate', 0.31) * 100 * f['view_rate'], 1)),
            (f'ctr_{mkt}_{ch}_{goal}_{sid}',              round(b['ctr'] * 100 * f['ctr'], 2)),
            (f'click_to_session_{mkt}_{ch}_{goal}_{sid}', round(b.get('click_to_session', 0.80) * 100, 0)),
            (f'conv_rate_{mkt}_{ch}_{goal}_{sid}',        round(b.get('conv_rate', 0.02) * 100 * f['conv_rate'], 1)),
        ]
    else:
        keys = [
            (f'cpm_{mkt}_{ch}_{goal}_{sid}',              round(b['cpm'] * f['cpm'], 2)),
            (f'ctr_{mkt}_{ch}_{goal}_{sid}',              round(b['ctr'] * 100 * f['ctr'], 2)),
            (f'click_to_session_{mkt}_{ch}_{goal}_{sid}', round(b.get('click_to_session', 0.82) * 100, 0)),
            (f'conv_rate_{mkt}_{ch}_{goal}_{sid}',        round(b.get('conv_rate', 0.02) * 100 * f['conv_rate'], 1)),
        ]
    for k, v in keys:
        st.session_state[k] = v


def _scenario_status(sid):
    ss = st.session_state
    has_goals   = any(ss.get(f'sb_goal_{g}_{sid}', False) for g in ALL_GOALS)
    has_markets = bool(ss.get(f'selected_markets_{sid}', []))
    has_budget  = ss.get(f'total_budget_{sid}', 0) > 0
    mkts        = ss.get(f'selected_markets_{sid}', [])
    pct_sum     = sum(ss.get(f'pct_{m}_{sid}', 0) for m in mkts)
    split_ok    = has_markets and abs(pct_sum - 100) <= 0.5
    steps = [
        ('Goals',   has_goals),
        ('Markets', has_markets),
        ('Budget',  has_budget),
        ('Split',   split_ok),
    ]
    parts = []
    for label, ok in steps:
        color = '#2BB5A5' if ok else '#E8A838'
        icon  = '✓' if ok else '○'
        parts.append(
            f'<span style="background:{color};color:white;border-radius:4px;'
            f'padding:2px 8px;font-size:0.72rem;font-weight:600;margin-right:4px">'
            f'{icon} {label}</span>'
        )
    st.markdown('<div style="margin-bottom:8px">' + ''.join(parts) + '</div>', unsafe_allow_html=True)


def _market_donut(market_pcts):
    labels = [f'{MARKET_LABELS[m]}' for m in market_pcts]
    values = list(market_pcts.values())
    colors = DONUT_PALETTE[:len(labels)]
    fig = go.Figure(go.Pie(
        labels=labels, values=values, hole=0.55,
        marker={'colors': colors, 'line': {'color': 'white', 'width': 2}},
        textinfo='percent', textfont={'size': 10},
        hovertemplate='%{label}: %{value:.1f}%<extra></extra>',
    ))
    fig.update_layout(
        margin={'t': 10, 'b': 10, 'l': 10, 'r': 10},
        height=200,
        paper_bgcolor='rgba(0,0,0,0)',
        showlegend=True,
        legend={'font': {'size': 12}, 'orientation': 'v', 'x': 1.02},
    )
    return fig


def _pacing_chart(periods, budget, sid):
    if not periods or budget <= 0:
        return
    total_days = sum(p['days'] for p in periods) or 1
    labels = [p['label'] for p in periods]
    values = [round(budget * p['days'] / total_days, 0) for p in periods]
    fig = go.Figure(go.Bar(
        x=labels, y=values,
        marker_color='#2BB5A5',
        text=[f'€{v:,.0f}' for v in values],
        textposition='outside',
        textfont={'size': 12, 'color': '#1A1A1A'},
    ))
    fig.update_layout(
        title={'text': 'Budget Pacing', 'font': {'size': 13, 'color': '#1A1A1A'}, 'x': 0.5, 'xanchor': 'center'},
        margin={'t': 40, 'b': 30, 'l': 10, 'r': 10},
        height=240,
        paper_bgcolor='rgba(0,0,0,0)',
        yaxis={'showticklabels': False, 'showgrid': False, 'zeroline': False},
        xaxis={'tickangle': -30 if len(periods) > 6 else 0, 'tickfont': {'size': 11}},
    )
    st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False}, key=f'pacing_{sid}')


def benchmark_inputs(ch, mkt, goal, sid=0):
    """Render benchmark inputs for one channel/market/goal. Returns a benchmark dict."""
    b = BENCH[mkt][ch]

    # Preset buttons
    pc = st.columns([1, 1, 1, 4])
    for i, pname in enumerate(['Conservative', 'Average', 'Aggressive']):
        if pc[i].button(pname, key=f'preset_{pname}_{ch}_{mkt}_{goal}_{sid}', use_container_width=True):
            _apply_bench_preset(ch, mkt, goal, sid, pname)
            st.rerun()

    fields = []
    if ch == 'Search':
        fields = [
            ('cpc', 'CPC (€)',   b['cpc'],              0.10, '%.2f', False),
            ('ctr', 'CTR %',     b['ctr'] * 100,        0.05, '%.2f', True),
        ]
        if goal in ('Traffic', 'Conversion'):
            fields.append(('click_to_session', 'Click→Session %',
                           b.get('click_to_session', 0.85) * 100, 1.0, '%.0f', True))
        if goal == 'Conversion':
            fields += [
                ('conv_rate',   'Session→Lead %', b.get('conv_rate',   0.03) * 100, 0.1, '%.1f', True),
                ('lead_to_mql', 'Lead→MQL %',     b.get('lead_to_mql', 0.20) * 100, 1.0, '%.0f', True),
                ('mql_to_sql',  'MQL→SQL %',      b.get('mql_to_sql',  0.30) * 100, 1.0, '%.0f', True),
            ]

    elif ch == 'YouTube':
        fields = [('cpm', 'CPM (€)', b['cpm'], 0.5, '%.2f', False)]
        if goal == 'Awareness':
            fields.append(('view_rate', 'View Rate %',
                           b.get('view_rate', 0.31) * 100, 0.5, '%.1f', True))
        fields += [
            ('ctr',       'CTR %',     b['ctr'] * 100,          0.01, '%.2f', True),
            ('frequency', 'Frequency', b.get('frequency', 3.0), 0.5,  '%.1f', False),
        ]
        if goal in ('Traffic', 'Conversion'):
            fields.append(('click_to_session', 'Click→Session %',
                           b.get('click_to_session', 0.80) * 100, 1.0, '%.0f', True))
        if goal == 'Conversion':
            fields += [
                ('conv_rate',   'Session→Lead %', b.get('conv_rate',   0.02) * 100, 0.1, '%.1f', True),
                ('lead_to_mql', 'Lead→MQL %',     b.get('lead_to_mql', 0.20) * 100, 1.0, '%.0f', True),
                ('mql_to_sql',  'MQL→SQL %',      b.get('mql_to_sql',  0.30) * 100, 1.0, '%.0f', True),
            ]

    else:  # LinkedIn
        fields = [
            ('cpm',       'CPM (€)',   b['cpm'],                0.5,  '%.2f', False),
            ('ctr',       'CTR %',     b['ctr'] * 100,          0.01, '%.2f', True),
            ('frequency', 'Frequency', b.get('frequency', 3.0), 0.5,  '%.1f', False),
        ]
        if goal in ('Traffic', 'Conversion'):
            fields.append(('click_to_session', 'Click→Session %',
                           b.get('click_to_session', 0.82) * 100, 1.0, '%.0f', True))
        if goal == 'Conversion':
            fields += [
                ('conv_rate',   'Session→Lead %', b.get('conv_rate',   0.02) * 100, 0.1, '%.1f', True),
                ('lead_to_mql', 'Lead→MQL %',     b.get('lead_to_mql', 0.20) * 100, 1.0, '%.0f', True),
                ('mql_to_sql',  'MQL→SQL %',      b.get('mql_to_sql',  0.30) * 100, 1.0, '%.0f', True),
            ]

    cols = st.columns(len(fields))
    raw = {}
    for i, (key, label, default, step, fmt, _) in enumerate(fields):
        raw[key] = cols[i].number_input(label, value=float(default), step=step, format=fmt,
                                        help=BENCH_HELP.get(key, ''),
                                        key=f'{key}_{mkt}_{ch}_{goal}_{sid}')

    bm = {}
    for key, _, _, _, _, is_pct in fields:
        bm[key] = raw[key] / 100.0 if is_pct else raw[key]

    # Carry over non-editable defaults
    if ch == 'YouTube' and goal != 'Awareness':
        bm['view_rate'] = b.get('view_rate', 0.31)
    if ch in ('YouTube', 'LinkedIn') and 'frequency' not in bm:
        bm['frequency'] = b.get('frequency', 3.0)
    if 'click_to_session' not in bm:
        bm['click_to_session'] = b.get('click_to_session', 0.80)
    if 'conv_rate' not in bm:
        bm['conv_rate'] = b.get('conv_rate', 0.02)
    if 'lead_to_mql' not in bm:
        bm['lead_to_mql'] = b.get('lead_to_mql', 0.20)
    if 'mql_to_sql' not in bm:
        bm['mql_to_sql'] = b.get('mql_to_sql', 0.30)

    return bm


def _channel_budget_split(mkt, goal, goal_chs, mkt_budget, sid=0):
    """Render per-channel budget split for one market/goal pair. Returns {ch: budget}."""
    if len(goal_chs) == 1:
        return {goal_chs[0]: mkt_budget}

    st.markdown('**Channel budget split**')

    if len(goal_chs) == 2:
        ch_a, ch_b = goal_chs
        pct_a = st.slider(
            f'{ch_a}  ←——→  {ch_b}',
            min_value=0, max_value=100, value=50, step=5,
            key=f'split_{mkt}_{goal}_{sid}',
            format='%d%%',
        )
        pct_b = 100 - pct_a
        bud_a = mkt_budget * pct_a / 100
        bud_b = mkt_budget * pct_b / 100
        # Show the resulting amounts side by side
        ca, cb = st.columns(2)
        ca.metric(ch_a, f'€{bud_a:,.0f}', f'{pct_a}%')
        cb.metric(ch_b, f'€{bud_b:,.0f}', f'{pct_b}%')
        return {ch_a: bud_a, ch_b: bud_b}

    # 3 channels: individual % inputs that should sum to 100
    default_pct = round(100 / len(goal_chs), 1)
    pcts = {}
    cols = st.columns(len(goal_chs))
    for i, ch in enumerate(goal_chs):
        pcts[ch] = cols[i].number_input(
            f'{ch} %', min_value=0.0, max_value=100.0,
            value=default_pct, step=5.0, format='%.0f',
            key=f'split_{mkt}_{goal}_{ch}_{sid}'
        )
    pct_sum = sum(pcts.values())
    ch_budgets = {ch: mkt_budget * pcts[ch] / pct_sum for ch in goal_chs} if pct_sum > 0 else {ch: 0 for ch in goal_chs}
    if abs(pct_sum - 100) > 0.5:
        st.caption(f'Percentages sum to {pct_sum:.0f}% — normalising proportionally.')
    amt_cols = st.columns(len(goal_chs))
    for i, ch in enumerate(goal_chs):
        amt_cols[i].metric(ch, f'€{ch_budgets[ch]:,.0f}')
    return ch_budgets


def render_goal_section(mkt, goal, selected_channels, ch_budgets, periods, grand_totals, sid=0):
    """Render benchmarks + table + funnel for one market/goal combo."""
    bm_all = {}
    with st.expander('Edit benchmarks', expanded=True):
        for ch in selected_channels:
            st.markdown(f'**{ch}**')
            bm_all[ch] = benchmark_inputs(ch, mkt, goal, sid)

    for ch in selected_channels:
        bm = bm_all[ch]
        ch_bud = ch_budgets[ch]
        conv_rate = bm.get('conv_rate', 0.02)

        df = build_table(periods, ch_bud, bm, goal, ch, conv_rate)

        col_t, col_f = st.columns([3, 2])
        with col_t:
            st.markdown(f'**{ch}** — €{ch_bud:,.0f}')
            st.dataframe(fmt_df(df, ch, goal), use_container_width=True, hide_index=True)
        with col_f:
            st.plotly_chart(
                make_funnel(df, goal, ch, f'{mkt} — {ch}'),
                use_container_width=True,
                config={'displayModeBar': False},
                key=f'funnel_{mkt}_{ch}_{goal}_{sid}',
            )

        t_row = df[df['Period'] == 'TOTAL'].copy()
        t_row['Market'] = MARKET_LABELS[mkt]
        grand_totals[goal][ch].append(t_row)


# ── Save / Load helpers ───────────────────────────────────────────────────────

_SKIP_KEYS = {
    '_pending_load', 'FormSubmitter', '_uploader_v',
    # button / download widget keys — never safe to restore
    'dup_', 'remove_', 'tpl_apply_', 'grp_', 'eq_', 'cpm_eff_',
    'pin_', 'dl_gads_', 'dl_excel', 'btn_', 'preset_',
    'save_tpl_', 'del_tpl_',
}

def _serialise_state():
    data = {'scenario_names': st.session_state.get('scenario_names', ['Scenario 1'])}
    # custom_templates is a dict of dicts — handle it explicitly
    if 'custom_templates' in st.session_state:
        data['custom_templates'] = st.session_state['custom_templates']
    for k, v in st.session_state.items():
        if any(k.startswith(s) for s in _SKIP_KEYS):
            continue
        if k == 'custom_templates':
            continue  # already handled above
        if isinstance(v, date):
            data[k] = {'__date__': v.isoformat()}
        elif isinstance(v, (str, int, float, bool, list)):
            data[k] = v
    return json.dumps(data, indent=2).encode('utf-8')


# ── Excel builder ─────────────────────────────────────────────────────────────

_NMQ_TEAL   = '2BB5A5'
_NMQ_BLACK  = '1A1A1A'
_NMQ_LIGHT  = 'F5F7F7'

def _xl_header(ws, row_num):
    fill = PatternFill('solid', fgColor=_NMQ_TEAL)
    font = Font(bold=True, color='FFFFFF', name='Calibri')
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center')


def _get_bm_ss(ch, mkt, goal, sid):
    """Read benchmark values from session state (mirrors benchmark_inputs without rendering)."""
    ss = st.session_state
    b  = BENCH[mkt][ch]

    def _pct(key, default):
        return ss.get(f'{key}_{mkt}_{ch}_{goal}_{sid}', default * 100) / 100

    bm = {}
    if ch == 'Search':
        bm['cpc']             = ss.get(f'cpc_{mkt}_{ch}_{goal}_{sid}', b['cpc'])
        bm['ctr']             = _pct('ctr',             b['ctr'])
        bm['click_to_session']= _pct('click_to_session',b.get('click_to_session', 0.85))
        bm['conv_rate']       = _pct('conv_rate',       b.get('conv_rate', 0.03))
    elif ch == 'YouTube':
        bm['cpm']             = ss.get(f'cpm_{mkt}_{ch}_{goal}_{sid}', b['cpm'])
        bm['ctr']             = _pct('ctr',             b['ctr'])
        bm['frequency']       = ss.get(f'frequency_{mkt}_{ch}_{goal}_{sid}', b.get('frequency', 3.0))
        bm['view_rate']       = (_pct('view_rate', b.get('view_rate', 0.31))
                                 if goal == 'Awareness' else b.get('view_rate', 0.31))
        bm['click_to_session']= _pct('click_to_session',b.get('click_to_session', 0.80))
        bm['conv_rate']       = _pct('conv_rate',       b.get('conv_rate', 0.02))
    else:  # LinkedIn
        bm['cpm']             = ss.get(f'cpm_{mkt}_{ch}_{goal}_{sid}', b['cpm'])
        bm['ctr']             = _pct('ctr',             b['ctr'])
        bm['frequency']       = ss.get(f'frequency_{mkt}_{ch}_{goal}_{sid}', b.get('frequency', 3.0))
        bm['click_to_session']= _pct('click_to_session',b.get('click_to_session', 0.82))
        bm['conv_rate']       = _pct('conv_rate',       b.get('conv_rate', 0.02))

    # MQL/SQL ratios — present for all channels when goal is Conversion
    bm['lead_to_mql'] = _pct('lead_to_mql', b.get('lead_to_mql', 0.20))
    bm['mql_to_sql']  = _pct('mql_to_sql',  b.get('mql_to_sql',  0.30))
    return bm


def _get_ch_budgets_ss(mkt, goal, goal_chs, mkt_budget, sid):
    """Read channel budget splits from session state without rendering UI."""
    ss = st.session_state
    if len(goal_chs) == 1:
        return {goal_chs[0]: mkt_budget}
    if len(goal_chs) == 2:
        ch_a, ch_b = goal_chs
        pct_a = ss.get(f'split_{mkt}_{goal}_{sid}', 50)
        return {ch_a: mkt_budget * pct_a / 100, ch_b: mkt_budget * (100 - pct_a) / 100}
    default_pct = round(100 / len(goal_chs), 1)
    pcts = {ch: ss.get(f'split_{mkt}_{goal}_{ch}_{sid}', default_pct) for ch in goal_chs}
    pct_sum = sum(pcts.values()) or 1
    return {ch: mkt_budget * pcts[ch] / pct_sum for ch in goal_chs}


def _build_excel_all(all_scenarios, scenario_ids, campaign_name, start_date, end_date, breakdown='Weekly'):
    """
    One workbook. One tab per country (× scenario if multiple).
    Layout: metric blocks (3 per horizontal row, periods as rows) + daily budget section below.
    """
    from openpyxl.utils import get_column_letter

    # ── Colour palette ────────────────────────────────────────────────────────
    C_DARK_BLUE  = PatternFill('solid', fgColor='1F3864')  # dark navy — block headers / title
    C_MED_BLUE   = PatternFill('solid', fgColor='2E75B6')  # medium blue — channel sub-headers
    C_LIGHT_BLUE = PatternFill('solid', fgColor='CFE1F3')  # light blue — TOTAL rows
    C_PERIOD     = PatternFill('solid', fgColor='D9E1F2')  # soft blue-grey — period label cells
    C_DATA       = PatternFill('solid', fgColor='FFFFFF')  # white — data cells
    C_DAILY      = PatternFill('solid', fgColor='FFF2CC')  # light yellow — daily budget cells

    F_WHITE_B = Font(color='FFFFFF', bold=True)
    F_BOLD    = Font(bold=True)
    F_NORMAL  = Font()
    ALIGN_C   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_L   = Alignment(horizontal='left', vertical='center')

    # Nine metric blocks arranged as three stacked sections (3 blocks each)
    SECTIONS = [
        [
            ('INVESTMENT',  'budget',       '#,##0.00'),
            ('CPC / CPM',   'cpc_cpm',      '#,##0.00'),
            ('IMPRESSIONS', 'impressions',  '#,##0'),
        ],
        [
            ('CTR',         'ctr',          '0.00%'),
            ('CLICKS',      'clicks',       '#,##0'),
            ('SESSIONS',    'sessions',     '#,##0'),
        ],
        [
            ('CONV. RATE',  'conv_rate',    '0.00%'),
            ('LEADS',       'conversions',  '#,##0'),
            ('CPL',         'cpl',          '#,##0.00'),
        ],
    ]
    CH_ORDER = ['LinkedIn', 'Search', 'YouTube']

    # ── Tab grouping: one per (market × scenario if market appears in 2+) ────
    market_count = {}
    for s in all_scenarios:
        for mkt in s['selected_markets']:
            market_count[mkt] = market_count.get(mkt, 0) + 1

    tabs = []
    for s, sid in zip(all_scenarios, scenario_ids):
        for mkt in s['selected_markets']:
            if market_count[mkt] > 1:
                tab_name = f"{MARKET_LABELS.get(mkt, mkt)} S{sid + 1}"[:31]
            else:
                tab_name = MARKET_LABELS.get(mkt, mkt)[:31]
            tabs.append((tab_name, s, sid, mkt))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    periods = generate_periods(start_date, end_date, breakdown)
    total_days = sum(p['days'] for p in periods) or 1

    for tab_name, s, sid, mkt in tabs:
        ws = wb.create_sheet(title=tab_name)
        mkt_label  = MARKET_LABELS.get(mkt, mkt)
        mkt_budget = s['market_budgets'][mkt]

        active_channels = [c for c in CH_ORDER
                           if any(c in chs for chs in s['goal_channels'].values())]
        n_ch = len(active_channels) or 1
        n_section_cols = 1 + 3 * n_ch  # period col + 3 blocks × n channels

        # ── Accumulate per-channel, per-period metrics across all goals ──────
        period_raw = {ch: {} for ch in active_channels}
        total_raw  = {ch: {'budget': 0, 'impressions': 0, 'clicks': 0,
                           'sessions': 0, 'conversions': 0}
                      for ch in active_channels}

        for goal, goal_chs in s['goal_channels'].items():
            ch_budgets = _get_ch_budgets_ss(mkt, goal, goal_chs, mkt_budget, sid)
            for ch in goal_chs:
                if ch not in active_channels:
                    continue
                ch_bud    = ch_budgets.get(ch, 0)
                bm        = _get_bm_ss(ch, mkt, goal, sid)
                conv_rate = bm.get('conv_rate', 0.02)

                for p in periods:
                    bud_p = ch_bud * p['days'] / total_days
                    row_m = calc_row(bud_p, bm, goal, ch, conv_rate)

                    if p['label'] not in period_raw[ch]:
                        period_raw[ch][p['label']] = {
                            'days': p['days'],
                            'budget': 0, 'impressions': 0, 'clicks': 0,
                            'sessions': 0, 'conversions': 0,
                        }
                    d = period_raw[ch][p['label']]
                    d['budget']      += row_m.get('Budget', 0)
                    d['impressions'] += row_m.get('impressions', 0)
                    d['clicks']      += row_m.get('clicks', 0)
                    d['sessions']    += row_m.get('sessions', 0)
                    d['conversions'] += row_m.get('conversions', 0)

                    t = total_raw[ch]
                    t['budget']      += row_m.get('Budget', 0)
                    t['impressions'] += row_m.get('impressions', 0)
                    t['clicks']      += row_m.get('clicks', 0)
                    t['sessions']    += row_m.get('sessions', 0)
                    t['conversions'] += row_m.get('conversions', 0)

        def _derive(ch, d):
            bud  = d.get('budget', 0)
            imp  = d.get('impressions', 0)
            clk  = d.get('clicks', 0)
            ses  = d.get('sessions', 0)
            conv = d.get('conversions', 0)
            cpc_cpm = (bud / clk if clk > 0 else 0) if ch == 'Search' \
                      else (bud / imp * 1000 if imp > 0 else 0)
            return {
                'budget':      bud,
                'impressions': imp,
                'clicks':      clk,
                'sessions':    ses,
                'conversions': conv,
                'cpc_cpm':     cpc_cpm,
                'ctr':         clk / imp if imp > 0 else 0,
                'conv_rate':   conv / ses if ses > 0 else 0,
                'cvr':         conv / clk if clk > 0 else 0,
                'cpl':         bud / conv if conv > 0 else 0,
            }

        # ── Title rows ───────────────────────────────────────────────────────
        row = 1
        flight = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
        title  = f"{mkt_label}  —  {campaign_name}  —  {s['name']}"
        for r_off, (txt, fnt, ht) in enumerate([
            (title, Font(color='FFFFFF', bold=True, size=13), 22),
            (f"Budget: €{mkt_budget:,.0f}  |  {flight}  |  {breakdown}", Font(color='FFFFFF', size=10), 15),
        ]):
            cell = ws.cell(row=row + r_off, column=1, value=txt)
            cell.fill = C_DARK_BLUE; cell.font = fnt; cell.alignment = ALIGN_L
            ws.merge_cells(start_row=row + r_off, start_column=1,
                           end_row=row + r_off, end_column=n_section_cols)
            ws.row_dimensions[row + r_off].height = ht
        row += 3  # 2 title rows + 1 blank

        # ── Metric sections (3 stacked, each with 3 blocks horizontally) ─────
        for section in SECTIONS:
            # Block-name header row (dark blue, merged per block)
            cell = ws.cell(row=row, column=1, value='')
            cell.fill = C_DARK_BLUE
            for b_idx, (block_name, metric_key, num_fmt) in enumerate(section):
                sc = 2 + b_idx * n_ch
                ec = sc + n_ch - 1
                cell = ws.cell(row=row, column=sc, value=block_name)
                cell.fill = C_DARK_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
                if n_ch > 1:
                    ws.merge_cells(start_row=row, start_column=sc,
                                   end_row=row, end_column=ec)
                for ci in range(sc + 1, ec + 1):
                    ws.cell(row=row, column=ci).fill = C_DARK_BLUE
            ws.row_dimensions[row].height = 18
            row += 1

            # Channel sub-header row (medium blue)
            cell = ws.cell(row=row, column=1, value='PERIOD')
            cell.fill = C_MED_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
            for b_idx in range(len(section)):
                for ch_idx, ch in enumerate(active_channels):
                    cell = ws.cell(row=row, column=2 + b_idx * n_ch + ch_idx, value=ch)
                    cell.fill = C_MED_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
            ws.row_dimensions[row].height = 16
            row += 1

            # Period data rows
            for p in periods:
                p_label = p['label']
                cell = ws.cell(row=row, column=1, value=p_label)
                cell.fill = C_PERIOD; cell.font = F_BOLD; cell.alignment = ALIGN_C
                for b_idx, (block_name, metric_key, num_fmt) in enumerate(section):
                    for ch_idx, ch in enumerate(active_channels):
                        col = 2 + b_idx * n_ch + ch_idx
                        derived = _derive(ch, period_raw[ch].get(p_label, {}))
                        val = derived.get(metric_key, 0)
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.fill = C_DATA; cell.font = F_NORMAL; cell.alignment = ALIGN_C
                        cell.number_format = num_fmt
                row += 1

            # TOTAL row
            cell = ws.cell(row=row, column=1, value='TOTAL')
            cell.fill = C_LIGHT_BLUE; cell.font = F_BOLD; cell.alignment = ALIGN_C
            for b_idx, (block_name, metric_key, num_fmt) in enumerate(section):
                for ch_idx, ch in enumerate(active_channels):
                    col = 2 + b_idx * n_ch + ch_idx
                    derived = _derive(ch, total_raw[ch])
                    val = derived.get(metric_key, 0)
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.fill = C_LIGHT_BLUE; cell.font = F_BOLD; cell.alignment = ALIGN_C
                    cell.number_format = num_fmt
            ws.row_dimensions[row].height = 15
            row += 2  # TOTAL + blank separator

        # ── Daily Budget section ──────────────────────────────────────────────
        row += 1
        db_n_cols = 2 + n_ch  # period + channels + total
        cell = ws.cell(row=row, column=1, value='DAILY BUDGET')
        cell.fill = C_DARK_BLUE; cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = ALIGN_L
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=db_n_cols)
        ws.row_dimensions[row].height = 18
        row += 1

        # Channel headers
        cell = ws.cell(row=row, column=1, value='PERIOD')
        cell.fill = C_MED_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
        for ch_idx, ch in enumerate(active_channels):
            cell = ws.cell(row=row, column=2 + ch_idx, value=ch)
            cell.fill = C_MED_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
        cell = ws.cell(row=row, column=2 + n_ch, value='TOTAL')
        cell.fill = C_MED_BLUE; cell.font = F_WHITE_B; cell.alignment = ALIGN_C
        ws.row_dimensions[row].height = 16
        row += 1

        for p in periods:
            p_label = p['label']
            cell = ws.cell(row=row, column=1, value=p_label)
            cell.fill = C_PERIOD; cell.font = F_BOLD; cell.alignment = ALIGN_C
            day_total = 0
            for ch_idx, ch in enumerate(active_channels):
                d    = period_raw[ch].get(p_label, {})
                days = d.get('days', p['days']) or 1
                daily = d.get('budget', 0) / days
                day_total += daily
                cell = ws.cell(row=row, column=2 + ch_idx, value=round(daily, 2))
                cell.fill = C_DAILY; cell.font = F_NORMAL; cell.alignment = ALIGN_C
                cell.number_format = '#,##0.00'
            cell = ws.cell(row=row, column=2 + n_ch, value=round(day_total, 2))
            cell.fill = C_DAILY; cell.font = F_BOLD; cell.alignment = ALIGN_C
            cell.number_format = '#,##0.00'
            row += 1

        # Daily budget TOTAL row (average daily across full flight)
        cell = ws.cell(row=row, column=1, value='TOTAL')
        cell.fill = C_LIGHT_BLUE; cell.font = F_BOLD; cell.alignment = ALIGN_C
        avg_total = 0
        for ch_idx, ch in enumerate(active_channels):
            avg = total_raw[ch]['budget'] / total_days
            avg_total += avg
            cell = ws.cell(row=row, column=2 + ch_idx, value=round(avg, 2))
            cell.fill = C_LIGHT_BLUE; cell.font = F_BOLD; cell.alignment = ALIGN_C
            cell.number_format = '#,##0.00'
        cell = ws.cell(row=row, column=2 + n_ch, value=round(avg_total, 2))
        cell.fill = C_LIGHT_BLUE; cell.font = F_BOLD; cell.alignment = ALIGN_C
        cell.number_format = '#,##0.00'

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 20
        for i in range(2, n_section_cols + 2):
            ws.column_dimensions[get_column_letter(i)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Google Ads CSV builder ────────────────────────────────────────────────────

_MARKET_LANGUAGE = {
    'AT': 'German', 'BE': 'French', 'BG': 'Bulgarian', 'HR': 'Croatian',
    'CY': 'Greek', 'CZ': 'Czech', 'DK': 'Danish', 'EE': 'Estonian',
    'FI': 'Finnish', 'FR': 'French', 'DE': 'German', 'GR': 'Greek',
    'HU': 'Hungarian', 'IE': 'English', 'IT': 'Italian', 'LV': 'Latvian',
    'LT': 'Lithuanian', 'LU': 'French', 'MT': 'English', 'NL': 'Dutch',
    'NO': 'Norwegian', 'PL': 'Polish', 'PT': 'Portuguese', 'RO': 'Romanian',
    'SK': 'Slovak', 'SI': 'Slovenian', 'ES': 'Spanish', 'SE': 'Swedish',
    'CH': 'German', 'UK': 'English',
}
_GADS_TYPE = {'YouTube': 'Video', 'Search': 'Search'}
_GADS_BID = {
    ('YouTube', 'Awareness'):  'Target CPM',
    ('YouTube', 'Traffic'):    'Maximize conversions',
    ('YouTube', 'Conversion'): 'Target CPA',
    ('Search',  'Awareness'):  'Manual CPC',
    ('Search',  'Traffic'):    'Maximize clicks',
    ('Search',  'Conversion'): 'Target CPA',
}
_GADS_COLS = [
    'Campaign', 'Campaign Status', 'Campaign Type',
    'Budget', 'Budget Type',
    'Bid Strategy Type', 'Target CPA', 'Target CPM',
    'Start Date', 'End Date',
    'Location', 'Location Type',
    'Language', 'Language Type',
    'Ad Group', 'Ad Group Status', 'Default Max. CPC',
]


def _build_gads_csv_scenario(s_data, sid, campaign_name, start_date, end_date):
    """Build a Google Ads Editor CSV directly from a rendered scenario's data."""
    ss   = st.session_state
    days = max((end_date - start_date).days + 1, 1)
    s_dt = start_date.strftime('%m/%d/%Y')
    e_dt = end_date.strftime('%m/%d/%Y')
    rows = []

    def _empty():
        return {c: '' for c in _GADS_COLS}

    for mkt in s_data['selected_markets']:
        mkt_budget = s_data['market_budgets'][mkt]
        for goal, channels in s_data['goal_channels'].items():
            gads_chs = [ch for ch in channels if ch in _GADS_TYPE]
            if not gads_chs:
                continue

            if len(gads_chs) == 1:
                ch_buds = {gads_chs[0]: mkt_budget}
            elif len(gads_chs) == 2:
                pct_a = ss.get(f'split_{mkt}_{goal}_{sid}', 50)
                ch_buds = {
                    gads_chs[0]: mkt_budget * pct_a / 100,
                    gads_chs[1]: mkt_budget * (100 - pct_a) / 100,
                }
            else:
                pcts = {ch: ss.get(f'split_{mkt}_{goal}_{ch}_{sid}', 100 / len(gads_chs)) for ch in gads_chs}
                total = sum(pcts.values()) or 1
                ch_buds = {ch: mkt_budget * pcts[ch] / total for ch in gads_chs}

            for ch in gads_chs:
                daily      = round(ch_buds[ch] / days, 2)
                name       = f'{campaign_name}_{mkt}_{goal}_{ch}'
                bid        = _GADS_BID.get((ch, goal), 'Manual CPC')
                target_cpm = ss.get(f'cpm_{mkt}_{ch}_{goal}_{sid}', '') if ch == 'YouTube' else ''
                default_cpc = ss.get(f'cpc_{mkt}_{ch}_{goal}_{sid}', '') if ch == 'Search' else ''

                r = _empty()
                r.update({'Campaign': name, 'Campaign Status': 'Enabled',
                          'Campaign Type': _GADS_TYPE[ch], 'Budget': daily,
                          'Budget Type': 'Daily', 'Bid Strategy Type': bid,
                          'Target CPM': target_cpm, 'Start Date': s_dt, 'End Date': e_dt})
                rows.append(r)

                r = _empty()
                r.update({'Campaign': name, 'Location': MARKET_LABELS[mkt], 'Location Type': 'Location'})
                rows.append(r)

                r = _empty()
                r.update({'Campaign': name, 'Language': _MARKET_LANGUAGE.get(mkt, 'English'), 'Language Type': 'Language'})
                rows.append(r)

                r = _empty()
                r.update({'Campaign': name, 'Ad Group': f'{name}_AdGroup_01',
                          'Ad Group Status': 'Enabled', 'Default Max. CPC': default_cpc})
                rows.append(r)

    buf = io.StringIO()
    pd.DataFrame(rows, columns=_GADS_COLS).to_csv(buf, index=False)
    return buf.getvalue().encode('utf-8')


# ── Step label helper ─────────────────────────────────────────────────────────
def _step(n, label):
    return (
        f'<p style="margin:6px 0 2px 0;font-weight:600;font-size:0.88rem;color:#1A1A1A">'
        f'<span style="background:#2BB5A5;color:white;border-radius:50%;'
        f'width:18px;height:18px;display:inline-flex;align-items:center;justify-content:center;'
        f'font-size:0.68rem;font-weight:700;margin-right:6px;flex-shrink:0">{n}</span>'
        f'{label}</p>'
    )


# Apply any pending plan load before widgets render (backup for app.py handler)
if '_pending_load' in st.session_state:
    _load_data = st.session_state.pop('_pending_load')
    for _k, _v in _load_data.items():
        if any(_k.startswith(s) for s in _SKIP_KEYS):
            continue
        if isinstance(_v, dict) and '__date__' in _v:
            st.session_state[_k] = date.fromisoformat(_v['__date__'])
        else:
            st.session_state[_k] = _v


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(_step(1, 'Campaign Name'), unsafe_allow_html=True)
    campaign_name = st.text_input('Campaign Name', value='Campaign — 2026',
                                  key='campaign_name', label_visibility='collapsed')

    st.markdown(_step(2, 'Audience Type'), unsafe_allow_html=True)
    audience_type = st.radio('Audience', ['B2B', 'B2C'], horizontal=True,
                             label_visibility='collapsed', key='audience_type')

    st.markdown(_step(3, 'Industry'), unsafe_allow_html=True)
    INDUSTRIES_SB = [
        'Logistics, Supply Chain & Transportation',
        'Industrial, Manufacturing & Materials',
        'Enterprise SaaS, Technology & Platforms',
        'Financial Services, Fintech & Insurance',
        'Healthcare, Pharma & Life Sciences',
        'Consumer Brands & Retail',
        'Mobility, Travel & Automotive',
        'Information Services & Professional Platforms',
    ]
    industry = st.selectbox('Industry', INDUSTRIES_SB, label_visibility='collapsed', key='industry')

    st.markdown('---')
    st.markdown(_step(4, 'Flight Dates'), unsafe_allow_html=True)
    col_s, col_e = st.columns(2)
    start_date = col_s.date_input('Start', value=date(2026, 5, 12), key='start_date')
    end_date   = col_e.date_input('End',   value=date(2026, 6, 29), key='end_date')

    if end_date <= start_date:
        st.error('End must be after Start.')
        st.stop()

    st.caption(f'{(end_date - start_date).days + 1} days total')

    st.markdown(_step(5, 'Period Breakdown'), unsafe_allow_html=True)
    breakdown = st.selectbox('Breakdown', ['Daily', 'Weekly', 'Bi-Weekly', 'Monthly'],
                             index=1, label_visibility='collapsed', key='breakdown')

    # ── Save / Load ───────────────────────────────────────────────────────────
    st.markdown('---')
    st.markdown('**💾 Save / Load Plan**')
    st.download_button(
        label='Download plan (.json)',
        data=_serialise_state(),
        file_name=f'{campaign_name.replace(" ", "_")}_plan.json',
        mime='application/json',
        use_container_width=True,
    )
    if '_uploader_v' not in st.session_state:
        st.session_state['_uploader_v'] = 0
    uploaded = st.file_uploader('Load plan (.json)', type='json',
                                label_visibility='collapsed',
                                key=f'_plan_uploader_{st.session_state["_uploader_v"]}')
    if uploaded is not None:
        try:
            payload = json.loads(uploaded.read().decode('utf-8'))
            st.session_state['_pending_load'] = payload
            st.session_state['_uploader_v'] += 1  # new key on next render = clears the uploader
            st.rerun()
        except Exception as e:
            st.error(f'Could not load plan: {e}')


# ── Auto-save warning ─────────────────────────────────────────────────────────
_components.html("""
<script>
window.addEventListener('beforeunload', function(e) {
    e.preventDefault();
    e.returnValue = 'Download your plan JSON from the sidebar before leaving — unsaved changes will be lost.';
});
</script>
""", height=0)

# ── Main ──────────────────────────────────────────────────────────────────────
st.markdown(
    f'<h2 style="margin:0 0 2px 0;font-family:Inter,sans-serif;font-size:1.5rem;color:#1A1A1A">'
    f'{campaign_name}</h2>',
    unsafe_allow_html=True,
)
st.caption(f'{start_date.strftime("%b %d, %Y")} – {end_date.strftime("%b %d, %Y")}  ·  {breakdown}  ·  {audience_type}  ·  {industry}')
st.divider()

# Scenario management via session state
if 'scenario_names' not in st.session_state:
    st.session_state.scenario_names = ['Scenario 1']
if 'custom_templates' not in st.session_state:
    st.session_state['custom_templates'] = {}

add_col, _ = st.columns([1, 6])
if add_col.button('＋ Add Scenario'):
    n = len(st.session_state.scenario_names) + 1
    st.session_state.scenario_names.append(f'Scenario {n}')

periods = generate_periods(start_date, end_date, breakdown)

_has_compare = len(st.session_state.scenario_names) >= 2
_tab_labels = (['⚖ Compare'] if _has_compare else []) + st.session_state.scenario_names
_all_tabs = st.tabs(_tab_labels)
compare_tab = _all_tabs[0] if _has_compare else None
scenario_tabs = _all_tabs[1:] if _has_compare else _all_tabs
all_scenarios_data = []  # collected for AI section and Compare tab


def _render_scenario(sid):
    """Render per-scenario config (goals, channels, markets, budget, split) then tables and funnels."""

    # ── Header: rename + duplicate + remove ───────────────────────────────────
    n_col, dup_col, rem_col = st.columns([6, 1, 1])
    new_name = n_col.text_input('Scenario name', value=st.session_state.scenario_names[sid],
                                key=f'rename_{sid}', label_visibility='collapsed',
                                placeholder='Scenario name…')
    if new_name and new_name != st.session_state.scenario_names[sid]:
        st.session_state.scenario_names[sid] = new_name
        st.rerun()
    if dup_col.button('⧉ Dup', key=f'dup_{sid}', use_container_width=True, help='Duplicate this scenario'):
        _duplicate_scenario(sid)
    if rem_col.button('✕ Remove', key=f'remove_{sid}', use_container_width=True,
                      disabled=len(st.session_state.scenario_names) <= 1,
                      help='Remove this scenario'):
        st.session_state.scenario_names.pop(sid)
        st.rerun()

    # ── Completion status ─────────────────────────────────────────────────────
    _scenario_status(sid)

    # ── Templates ─────────────────────────────────────────────────────────────
    with st.expander('📁 Templates', expanded=False):
        custom_tpls = st.session_state.get('custom_templates', {})

        # ── Built-in ──────────────────────────────────────────────────────────
        st.markdown('<small style="color:#6b7280;font-weight:600;text-transform:uppercase;'
                    'letter-spacing:0.06em">Built-in</small>', unsafe_allow_html=True)
        tpl_options = ['— pick a template —'] + list(PLAN_TEMPLATES.keys())
        tpl_sel = st.selectbox('Template', tpl_options, key=f'tpl_{sid}',
                               label_visibility='collapsed')
        if tpl_sel != '— pick a template —':
            if st.button(f'Apply "{tpl_sel}"', key=f'tpl_apply_{sid}'):
                _apply_template(sid, tpl_sel)

        st.divider()

        # ── Your saved templates ──────────────────────────────────────────────
        st.markdown('<small style="color:#6b7280;font-weight:600;text-transform:uppercase;'
                    'letter-spacing:0.06em">Your saved templates</small>', unsafe_allow_html=True)
        if custom_tpls:
            for tpl_name, tpl_data in list(custom_tpls.items()):
                goals_summary = '  ·  '.join(
                    f"{g}: {', '.join(chs)}" for g, chs in tpl_data.get('goals', {}).items()
                )
                mkts = ', '.join(MARKET_LABELS.get(m, m) for m in tpl_data.get('markets', []))
                st.markdown(
                    f'<div style="font-size:0.82rem;font-weight:600;margin-bottom:2px">{tpl_name}</div>'
                    f'<div style="font-size:0.75rem;color:#6b7280;margin-bottom:6px">'
                    f'€{tpl_data.get("budget", 0):,.0f} · {mkts}<br>{goals_summary}</div>',
                    unsafe_allow_html=True,
                )
                ca, cb = st.columns([3, 1])
                safe = tpl_name.replace(' ', '_')[:24]
                if ca.button(f'Apply', key=f'tpl_apply_custom_{safe}_{sid}',
                             use_container_width=True):
                    _apply_template_data(sid, tpl_data)
                if cb.button('✕ Delete', key=f'del_tpl_{safe}_{sid}',
                             use_container_width=True):
                    del st.session_state['custom_templates'][tpl_name]
                    st.rerun()
        else:
            st.caption('No saved templates yet.')

        st.divider()

        # ── Save current as template ──────────────────────────────────────────
        st.markdown('<small style="color:#6b7280;font-weight:600;text-transform:uppercase;'
                    'letter-spacing:0.06em">Save current as template</small>',
                    unsafe_allow_html=True)
        new_tpl_name = st.text_input('Template name', key=f'tpl_name_{sid}',
                                     placeholder='e.g. Q3 DACH Full Funnel',
                                     label_visibility='collapsed')
        if st.button('💾 Save template', key=f'save_tpl_{sid}',
                     use_container_width=True, disabled=not new_tpl_name.strip()):
            name = new_tpl_name.strip()
            st.session_state['custom_templates'][name] = _current_as_template(sid)
            st.session_state[f'tpl_name_{sid}'] = ''
            st.success(f'"{name}" saved.')
            st.rerun()

    st.divider()

    # ── Goals & Channels ─────────────────────────────────────────────────────
    st.markdown('**Goals & Channels**')
    hc = st.columns([2, 1, 1, 1])
    hc[1].markdown('<small>YT</small>', unsafe_allow_html=True)
    hc[2].markdown('<small>Search</small>', unsafe_allow_html=True)
    hc[3].markdown('<small>LinkedIn</small>', unsafe_allow_html=True)

    goal_channels = {}
    for goal_key in ALL_GOALS:
        rc = st.columns([2, 1, 1, 1])
        goal_on = rc[0].checkbox(goal_key, value=False, key=f'sb_goal_{goal_key}_{sid}')
        yt_on = rc[1].checkbox('YT', value=False, key=f'sb_yt_{goal_key}_{sid}', label_visibility='collapsed', disabled=not goal_on)
        s_on  = rc[2].checkbox('S',  value=False, key=f'sb_s_{goal_key}_{sid}',  label_visibility='collapsed', disabled=not goal_on)
        li_on = rc[3].checkbox('LI', value=False, key=f'sb_li_{goal_key}_{sid}', label_visibility='collapsed', disabled=not goal_on)
        if goal_on:
            chs = [ch for ch, on in [('YouTube', yt_on), ('Search', s_on), ('LinkedIn', li_on)] if on]
            if chs:
                goal_channels[goal_key] = chs

    if not goal_channels:
        st.info('Select at least one goal and channel above.')
        return None

    selected_goals = list(goal_channels.keys())
    st.divider()

    # ── Markets & Budget ─────────────────────────────────────────────────────
    cfg1, cfg2 = st.columns([4, 1])
    with cfg1:
        st.markdown('**Markets**')
        # Market group shortcuts
        grp_cols = st.columns(len(MARKET_GROUPS))
        for gi, (grp_label, grp_mkts) in enumerate(MARKET_GROUPS.items()):
            if grp_cols[gi].button(grp_label, key=f'grp_{grp_label}_{sid}',
                                   use_container_width=True, help=f'Add {", ".join(grp_mkts)}'):
                current = list(st.session_state.get(f'selected_markets_{sid}', []))
                merged  = current + [m for m in grp_mkts if m not in current]
                st.session_state[f'selected_markets_{sid}'] = merged
                st.rerun()
        s_markets = st.multiselect(
            'Markets', list(MARKET_LABELS.keys()), default=[],
            format_func=lambda k: f'{k} — {MARKET_LABELS[k]}',
            label_visibility='collapsed',
            key=f'selected_markets_{sid}',
        )
    with cfg2:
        st.markdown('**Total Budget (€)**')
        s_budget = st.number_input(
            'Budget', min_value=0, value=0, step=500,
            label_visibility='collapsed', key=f'total_budget_{sid}'
        )

    s_market_budgets, s_market_pcts = {}, {}
    if s_markets:
        n_mkts = len(s_markets)
        default_pct = round(100.0 / n_mkts, 1)

        # Budget split shortcuts
        sc1, sc2, _ = st.columns([1, 1, 3])
        if sc1.button('⚖ Equal split', key=f'eq_{sid}', use_container_width=True):
            for m in s_markets:
                st.session_state[f'pct_{m}_{sid}'] = default_pct
            st.rerun()
        if sc2.button('📊 CPM-efficient', key=f'cpm_eff_{sid}', use_container_width=True,
                      help='More budget to cheaper markets to maximise impressions'):
            weights = {}
            for m in s_markets:
                cpms = [BENCH[m][ch]['cpm'] for ch in ['YouTube', 'LinkedIn']
                        if ch in BENCH[m] and 'cpm' in BENCH[m][ch]]
                weights[m] = 1.0 / (sum(cpms) / len(cpms)) if cpms else 0.1
            total_w = sum(weights.values()) or 1
            for m in s_markets:
                st.session_state[f'pct_{m}_{sid}'] = round(weights[m] / total_w * 100, 1)
            st.rerun()

        # Split inputs + donut side by side
        split_col, donut_col = st.columns([3, 2])
        with split_col:
            st.markdown('**Market Split (%)**')
            per_row = min(n_mkts, 3)
            for row_start in range(0, n_mkts, per_row):
                row_mkts = s_markets[row_start:row_start + per_row]
                row_cols = st.columns(len(row_mkts) * 2)
                for i, mkt in enumerate(row_mkts):
                    pct = row_cols[i * 2].number_input(
                        f'{mkt}', min_value=0.0, max_value=100.0,
                        value=default_pct, step=0.5, format='%.1f',
                        key=f'pct_{mkt}_{sid}'
                    )
                    s_market_pcts[mkt] = pct
                    s_market_budgets[mkt] = s_budget * pct / 100
                    row_cols[i * 2 + 1].markdown(
                        f'<div style="font-size:0.78rem;color:#555;margin-top:4px">{mkt}</div>'
                        f'<div style="font-size:0.9rem;font-weight:600;color:#1A1A1A">'
                        f'€{s_market_budgets[mkt]:,.0f}</div>',
                        unsafe_allow_html=True,
                    )
            pct_sum = sum(s_market_pcts.values())
            if abs(pct_sum - 100) > 0.5:
                st.warning(f'Split: {pct_sum:.1f}% — adjust to 100%')
            else:
                st.success(f'✓ €{s_budget:,} allocated')
        with donut_col:
            if s_market_pcts:
                st.plotly_chart(_market_donut(s_market_pcts), use_container_width=True,
                                config={'displayModeBar': False}, key=f'donut_{sid}')
    else:
        st.info('Select at least one market above to build the plan.')
        return None

    # Pacing chart (inline, compact)
    _pacing_chart(periods, s_budget, sid)

    st.divider()

    # ── Pinned country panel ──────────────────────────────────────────────────
    pinned_mkt = st.session_state.get(f'pinned_country_{sid}')
    if pinned_mkt and pinned_mkt in s_markets:
        pm_budget = s_market_budgets.get(pinned_mkt, 0)
        pm_pct    = s_market_pcts.get(pinned_mkt, 0)
        pm_goals  = '  ·  '.join(
            f'{g}: {", ".join(chs)}' for g, chs in goal_channels.items()
        )
        cached    = st.session_state.get(f'cached_kpis_{pinned_mkt}_{sid}', {})
        kpi_html  = ('  <span style="color:#666">|</span>  '.join(
            f'<b>{k}</b>: {v}' for k, v in cached.items()
        )) if cached else '<span style="color:#888;font-style:italic">KPIs appear after first render</span>'
        st.markdown(
            '<style>'
            '.pinned-bar{position:sticky;top:3.6rem;z-index:200;background:#f0faf9;'
            'border-left:4px solid #2BB5A5;border-radius:0 6px 6px 0;'
            'padding:7px 14px;margin-bottom:10px;box-shadow:0 2px 8px rgba(0,0,0,0.08);'
            'font-size:0.82rem;line-height:1.5}'
            '</style>'
            f'<div class="pinned-bar">📌 <strong>{MARKET_LABELS[pinned_mkt]}</strong>'
            f'&nbsp;—&nbsp;€{pm_budget:,.0f} ({pm_pct:.0f}%)'
            f'&nbsp;&nbsp;<span style="color:#2BB5A5;font-size:0.75rem">{pm_goals}</span>'
            f'<br>{kpi_html}</div>',
            unsafe_allow_html=True,
        )

    # ── Plan tables ───────────────────────────────────────────────────────────
    grand_totals = {g: {ch: [] for ch in chs} for g, chs in goal_channels.items()}

    for mkt in s_markets:
        mkt_budget = s_market_budgets[mkt]

        # Country heading with pin toggle
        h_col, pin_col = st.columns([12, 1])
        is_pinned = (st.session_state.get(f'pinned_country_{sid}') == mkt)
        h_col.subheader(MARKET_LABELS[mkt])
        pin_icon = '☑' if is_pinned else '☐'
        pin_tip  = 'Unpin this country' if is_pinned else 'Pin to top for comparison while scrolling'
        if pin_col.button(pin_icon, key=f'pin_{mkt}_{sid}',
                          use_container_width=True, help=pin_tip):
            if is_pinned:
                st.session_state.pop(f'pinned_country_{sid}', None)
            else:
                st.session_state[f'pinned_country_{sid}'] = mkt
            st.rerun()

        if len(selected_goals) > 1:
            goal_tabs = st.tabs(selected_goals)
            for tab, goal in zip(goal_tabs, selected_goals):
                with tab:
                    goal_chs = goal_channels[goal]
                    ch_budgets = _channel_budget_split(mkt, goal, goal_chs, mkt_budget, sid)
                    render_goal_section(mkt, goal, goal_chs, ch_budgets, periods, grand_totals, sid)
        else:
            goal = selected_goals[0]
            goal_chs = goal_channels[goal]
            ch_budgets = _channel_budget_split(mkt, goal, goal_chs, mkt_budget, sid)
            render_goal_section(mkt, goal, goal_chs, ch_budgets, periods, grand_totals, sid)

        # Cache KPI totals for the pinned panel (available on next render)
        if mkt == st.session_state.get(f'pinned_country_{sid}'):
            kpi_cache = {}
            for g in selected_goals:
                for ch in goal_channels[g]:
                    rows = grand_totals[g][ch]
                    if rows:
                        t = rows[-1].iloc[0]
                        parts = [
                            f'{COL_FMT[c][0]}: {COL_FMT[c][1](t[c])}'
                            for c in ['impressions', 'reach', 'views', 'clicks', 'sessions', 'conversions']
                            if c in t.index and t[c] > 0
                        ]
                        if parts:
                            kpi_cache[f'{g} · {ch}'] = ' | '.join(parts[:3])
            st.session_state[f'cached_kpis_{pinned_mkt}_{sid}'] = kpi_cache

        st.divider()

    st.subheader('Grand Total — All Markets')
    for goal in selected_goals:
        if len(selected_goals) > 1:
            st.markdown(f'### {goal}')
        for ch in goal_channels[goal]:
            rows = grand_totals[goal][ch]
            if not rows:
                continue
            gdf = pd.concat(rows, ignore_index=True)
            st.markdown(f'**{ch} — per market**')
            disp = gdf[['Market']].copy()
            for col in ADDITIVE:
                if col in gdf.columns and col in COL_FMT:
                    label, fn = COL_FMT[col]
                    disp[label] = gdf[col].apply(fn)
            st.dataframe(disp, use_container_width=True, hide_index=True)
            grand = {'Period': 'TOTAL', 'Days': int(gdf['Days'].sum())}
            for col in ADDITIVE:
                if col in gdf.columns:
                    grand[col] = gdf[col].sum()
            grand_df = pd.DataFrame([grand])
            g1, g2 = st.columns([2, 3])
            with g1:
                st.markdown(f'**{ch} — combined totals**')
                rows_disp = [{'Metric': COL_FMT[c][0], 'Value': COL_FMT[c][1](grand[c])}
                             for c in ADDITIVE if c in grand and c in COL_FMT]
                st.dataframe(pd.DataFrame(rows_disp), use_container_width=True, hide_index=True)
            with g2:
                st.plotly_chart(
                    make_funnel(grand_df, goal, ch, f'Grand Total — {ch} ({goal})'),
                    use_container_width=True, config={'displayModeBar': False},
                    key=f'grand_{goal}_{ch}_{sid}',
                )
        if len(selected_goals) > 1:
            st.divider()

    return {
        'name': st.session_state.scenario_names[sid],
        'grand_totals': grand_totals,
        'goal_channels': goal_channels,
        'selected_markets': s_markets,
        'market_budgets': s_market_budgets,
        'market_pcts': s_market_pcts,
        'grand_total_bud': s_budget,
    }


_scenario_sids = []
for sid, s_tab in enumerate(scenario_tabs):
    with s_tab:
        s_data = _render_scenario(sid)
        if s_data:
            all_scenarios_data.append(s_data)
            _scenario_sids.append(sid)
            st.divider()
            gads_bytes = _build_gads_csv_scenario(s_data, sid, campaign_name, start_date, end_date)
            st.download_button(
                label=f'⬇ Google Ads CSV — {s_data["name"]}',
                data=gads_bytes,
                file_name=f'{campaign_name.replace(" ", "_")}_{s_data["name"].replace(" ", "_")}_google_ads.csv',
                mime='text/csv',
                key=f'dl_gads_{sid}',
                use_container_width=True,
            )

# ── Combined Excel download (all scenarios as tabs) ───────────────────────────
if all_scenarios_data:
    xl_bytes = _build_excel_all(all_scenarios_data, _scenario_sids, campaign_name, start_date, end_date, breakdown)
    st.download_button(
        label=f'⬇ Download Excel — All Scenarios ({len(all_scenarios_data)} tab{"s" if len(all_scenarios_data) != 1 else ""})',
        data=xl_bytes,
        file_name=f'{campaign_name.replace(" ", "_")}_media_plan.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key='dl_excel_all',
        use_container_width=True,
    )


# ── Compare tab (only when 2+ scenarios) ─────────────────────────────────────

def get_api_key():
    # Streamlit Cloud: secrets injected via dashboard
    try:
        key = st.secrets['anthropic'].get('api_key') or st.secrets['anthropic'].get('ANTHROPIC_API_KEY', '')
        if key:
            return key
    except Exception:
        pass
    # Local fallback: read from .streamlit/secrets.toml next to app file
    local = os.path.join(os.path.dirname(__file__), '.streamlit', 'secrets.toml')
    try:
        s = toml.load(local)
        return s['anthropic'].get('api_key') or s['anthropic'].get('ANTHROPIC_API_KEY', '')
    except Exception:
        return ''

def _aggregate_scenario_metrics(s):
    """Sum all additive metrics across goals and channels for a scenario."""
    totals = {col: 0 for col in ADDITIVE}
    for goal, chs in s['grand_totals'].items():
        for ch, rows in chs.items():
            if not rows:
                continue
            gdf = pd.concat(rows, ignore_index=True)
            for col in ADDITIVE:
                if col in gdf.columns:
                    totals[col] += gdf[col].sum()
    return totals


def _kpi_rows(s, metrics):
    """Build text lines per goal+channel for the AI prompt."""
    lines = []
    for goal, chs in s['grand_totals'].items():
        for ch, rows in chs.items():
            if not rows:
                continue
            gdf = pd.concat(rows, ignore_index=True)
            parts = []
            for col in ADDITIVE:
                if col in gdf.columns and gdf[col].sum() > 0:
                    parts.append(f'{COL_FMT[col][0]}: {COL_FMT[col][1](gdf[col].sum())}')
            if parts:
                lines.append(f'    [{goal} / {ch}] {" | ".join(parts)}')
    return '\n'.join(lines)


if compare_tab is not None:
    with compare_tab:
        st.caption('Side-by-side KPI comparison and AI recommendation — budget efficiency, reach, and funnel performance.')
        if len(all_scenarios_data) < 2:
            st.info('Fill in at least two scenarios (markets, goals, channels, and budgets) to run a comparison.')
        else:
            # ── KPI summary table (always visible, no button needed) ──────────
            st.markdown('#### KPI Summary')
            metric_labels = {col: COL_FMT[col][0] for col in ADDITIVE if col in COL_FMT}
            summary_rows = []
            for s in all_scenarios_data:
                agg = _aggregate_scenario_metrics(s)
                row = {'Scenario': s['name'], 'Budget (€)': f"€{agg['Budget']:,.0f}"}
                for col in ADDITIVE:
                    if col == 'Budget':
                        continue
                    if agg.get(col, 0) > 0:
                        row[metric_labels[col]] = COL_FMT[col][1](agg[col])
                summary_rows.append(row)
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

            # ── Winner callouts per metric ────────────────────────────────────
            st.markdown('#### Which scenario leads on each KPI?')
            winner_cols = st.columns(len(ADDITIVE) - 1)  # skip Budget
            col_idx = 0
            for col in ADDITIVE:
                if col == 'Budget':
                    continue
                vals = [(s['name'], _aggregate_scenario_metrics(s).get(col, 0)) for s in all_scenarios_data]
                best = max(vals, key=lambda x: x[1])
                if best[1] > 0:
                    winner_cols[col_idx].metric(metric_labels[col], best[0], help=f'Highest {metric_labels[col]} across all scenarios')
                col_idx += 1

            # ── Visual comparison bar charts ──────────────────────────────────
            st.markdown('#### Visual comparison')
            bar_metrics = [c for c in ['impressions','reach','clicks','sessions','conversions']
                           if any(_aggregate_scenario_metrics(s).get(c, 0) > 0 for s in all_scenarios_data)]
            if bar_metrics:
                bar_cols = st.columns(min(len(bar_metrics), 3))
                sc_names  = [s['name'] for s in all_scenarios_data]
                sc_colors = DONUT_PALETTE[:len(all_scenarios_data)]
                for bi, metric in enumerate(bar_metrics[:6]):
                    with bar_cols[bi % 3]:
                        vals = [_aggregate_scenario_metrics(s).get(metric, 0) for s in all_scenarios_data]
                        fig  = go.Figure(go.Bar(
                            x=sc_names, y=vals,
                            marker_color=sc_colors,
                            text=[COL_FMT[metric][1](v) for v in vals],
                            textposition='outside',
                            textfont={'size': 9},
                        ))
                        fig.update_layout(
                            title={'text': COL_FMT[metric][0], 'font': {'size': 11}, 'x': 0.5, 'xanchor': 'center'},
                            margin={'t': 34, 'b': 10, 'l': 10, 'r': 10},
                            height=190,
                            paper_bgcolor='rgba(0,0,0,0)',
                            yaxis={'showticklabels': False, 'showgrid': False, 'zeroline': False},
                            xaxis={'tickfont': {'size': 9}},
                        )
                        st.plotly_chart(fig, use_container_width=True,
                                        config={'displayModeBar': False}, key=f'bar_{metric}')

            st.divider()

            # ── AI deep-dive ──────────────────────────────────────────────────
            if st.button('Generate AI Comparison', key='btn_compare'):
                api_key = get_api_key()
                if not api_key:
                    st.error('No API key found in .streamlit/secrets.toml')
                else:
                    def _scenario_block(s):
                        agg = _aggregate_scenario_metrics(s)
                        lines = [f"Scenario: {s['name']}"]
                        lines.append(f"  Total budget: €{s['grand_total_bud']:,.0f}")
                        lines.append(f"  Markets ({len(s['selected_markets'])}): {', '.join(MARKET_LABELS[m] for m in s['selected_markets'])}")
                        for goal, chs in s['goal_channels'].items():
                            lines.append(f"  Goal: {goal} | Channels: {', '.join(chs)}")
                        lines.append('  Aggregated KPIs:')
                        for col in ADDITIVE:
                            if col in agg and agg[col] > 0:
                                lines.append(f'    {COL_FMT[col][0]}: {COL_FMT[col][1](agg[col])}')
                        lines.append('  KPIs by goal and channel:')
                        lines.append(_kpi_rows(s, agg))
                        return '\n'.join(lines)

                    scenarios_text = '\n\n'.join(_scenario_block(s) for s in all_scenarios_data)
                    compare_prompt = f"""You are a senior paid media strategist comparing {len(all_scenarios_data)} media plan scenarios for a {audience_type} brand in the {industry} sector.

{scenarios_text}

For each scenario write a clearly labelled block using exactly this format (one block per scenario):

SCENARIO: [name]
STRENGTHS: [60–80 words — what this scenario does well for a {audience_type} {industry} brand: market coverage, channel fit, which KPIs it leads on and why that matters for this industry]
WEAKNESSES: [60–80 words — where this scenario underperforms: KPIs it loses on, missing markets, channel imbalance, or poor fit with {audience_type} buyer journeys in {industry}]
VERDICT: [25–35 words — one direct sentence on the single best use case for this scenario]

After all scenario blocks, write these two final sections:

BEST FOR BUDGET EFFICIENCY:
[50–60 words — which scenario delivers the most value per euro spent, referencing the actual KPI numbers. Frame this for a {audience_type} {industry} brand where budget scrutiny is typical.]

BEST FOR KPI PERFORMANCE:
[50–60 words — which scenario wins on raw KPI delivery per funnel stage (awareness → traffic → conversion), and what that means for a {audience_type} {industry} campaign objective.]

No bullet points. Write like a strategist presenting to a client."""

                    import anthropic as _anthropic
                    client = _anthropic.Anthropic(api_key=api_key)
                    with st.spinner('Comparing scenarios...'):
                        msg = client.messages.create(
                            model='claude-haiku-4-5-20251001',
                            max_tokens=1100,
                            messages=[{'role': 'user', 'content': compare_prompt}]
                        )
                    raw = msg.content[0].text
                    import re as _re
                    blocks = _re.split(r'\n(SCENARIO|STRENGTHS|WEAKNESSES|VERDICT|BEST FOR BUDGET EFFICIENCY|BEST FOR KPI PERFORMANCE):\s*', raw)
                    if len(blocks) > 1:
                        labels = blocks[1::2]
                        bodies = blocks[2::2]
                        palette = {
                            'SCENARIO':                   ('#1F497D', 'white'),
                            'STRENGTHS':                  ('#1F6152', 'white'),
                            'WEAKNESSES':                 ('#8B3A3A', 'white'),
                            'VERDICT':                    ('#437CA3', 'white'),
                            'BEST FOR BUDGET EFFICIENCY': ('#5A3E7A', 'white'),
                            'BEST FOR KPI PERFORMANCE':   ('#2D6A8A', 'white'),
                        }
                        for label, body in zip(labels, bodies):
                            bg, fg = palette.get(label, ('#555', 'white'))
                            st.markdown(
                                f'<div style="background:{bg};color:{fg};padding:6px 12px;border-radius:4px 4px 0 0;'
                                f'font-weight:bold;margin-top:10px">{label}</div>'
                                f'<div style="background:#f5f5f5;padding:10px 12px;border-radius:0 0 4px 4px;'
                                f'margin-bottom:2px">{body.strip()}</div>',
                                unsafe_allow_html=True,
                            )
                    else:
                        st.markdown(raw)


# ── AI Insights & Recommendations ────────────────────────────────────────────
st.divider()
st.subheader('AI Insights & Recommendations')

if all_scenarios_data:
    scenario_options = [s['name'] for s in all_scenarios_data]
    ai_scenario_name = st.selectbox('Scenario to analyse', scenario_options)
    ai_data = next(s for s in all_scenarios_data if s['name'] == ai_scenario_name)
else:
    ai_data = None

st.caption(f'Insights framed for **{audience_type}** · **{industry}**')

def build_plan_summary():
    """Compile a plain-text plan summary from the selected scenario's data."""
    if not ai_data:
        return 'No scenario data available.'
    d = ai_data
    lines = [
        f'Campaign: {campaign_name}  ({d["name"]})',
        f'Audience: {audience_type}  |  Industry: {industry}',
        f'Flight: {start_date.strftime("%b %d, %Y")} – {end_date.strftime("%b %d, %Y")} ({(end_date - start_date).days + 1} days)',
        f'Total budget: €{d["grand_total_bud"]:,}',
        f'Markets: {", ".join(MARKET_LABELS[m] for m in d["selected_markets"])}',
        '',
    ]
    for goal, chs in d['goal_channels'].items():
        lines.append(f'Goal: {goal}  |  Channels: {", ".join(chs)}')
        for ch in chs:
            rows = d['grand_totals'][goal][ch]
            if not rows:
                continue
            gdf = pd.concat(rows, ignore_index=True)
            grand = {col: gdf[col].sum() for col in ADDITIVE if col in gdf.columns}
            parts = [f'{COL_FMT[col][0]}: {COL_FMT[col][1](grand[col])}'
                     for col in ADDITIVE if col in grand and col in COL_FMT]
            lines.append(f'  {ch}: {" | ".join(parts)}')
        lines.append('')
    return '\n'.join(lines)

ai_tab_insights, ai_tab_recs, ai_tab_benchmarks = st.tabs([
    'Plan Insights', 'Market Recommendations', 'Benchmark Explanations'
])

with ai_tab_insights:
    st.caption('High-level read of the plan — what the numbers say, what looks strong, what to watch.')
    if st.button('Generate Plan Insights', key='btn_insights'):
        api_key = get_api_key()
        if not api_key:
            st.error('No API key found in .streamlit/secrets.toml')
        else:
            summary = build_plan_summary()
            prompt = f"""You are a senior paid media strategist reviewing a digital media plan for a {audience_type} brand in the {industry} sector.

{summary}

Write a concise but sharp analysis (150–200 words). Cover:
1. Overall scale and reach potential given the budget and markets — frame this for {audience_type} {industry} audiences specifically
2. Channel mix strengths or gaps for the selected goals — comment on what works for {audience_type} in this sector
3. Any markets where the allocation looks strong or under-invested for this industry
4. One clear watch-out or risk relevant to {audience_type} {industry} campaigns

Be direct. No headers. No bullet points. Write in plain paragraphs like a strategist talking to a client."""
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            with st.spinner('Thinking...'):
                msg = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=400,
                    messages=[{'role': 'user', 'content': prompt}]
                )
            st.markdown(msg.content[0].text)

with ai_tab_recs:
    st.caption('Budget allocation recommendations — which markets or channels deserve more weight and why.')
    if st.button('Generate Market Recommendations', key='btn_recs'):
        api_key = get_api_key()
        if not api_key:
            st.error('No API key found in .streamlit/secrets.toml')
        else:
            summary = build_plan_summary()
            mkt_list = '\n'.join(
                f'- {MARKET_LABELS[m]}: {ai_data["market_pcts"][m]:.1f}% (€{ai_data["market_budgets"][m]:,.0f})'
                for m in ai_data['selected_markets']
            ) if ai_data else 'No data.'
            channels_in_use = sorted({ch for chs in (ai_data['goal_channels'].values() if ai_data else []) for ch in chs})
            prompt = f"""You are a senior paid media strategist advising a {audience_type} brand in the {industry} sector on market budget allocation across {', '.join(channels_in_use) if channels_in_use else 'digital channels'}.

{summary}

Current market split:
{mkt_list}

Write exactly 3 recommendation sections, each starting with its label on its own line:

CURRENT ALLOCATION:
[50–60 words assessing whether the current market split makes sense for a {audience_type} {industry} brand — consider CPM efficiency vs audience quality for this sector and whether the channels in use ({', '.join(channels_in_use) if channels_in_use else 'the selected channels'}) justify the current weights]

REBALANCING OPTION:
[50–60 words suggesting one concrete rebalancing move — e.g. shift 10% from X to Y — grounded in what works for {audience_type} audiences in {industry} and the channel mix being used]

BEST ALLOCATION:
[50–60 words giving a direct final recommendation on the optimal split, framed specifically for a {audience_type} {industry} brand running {', '.join(channels_in_use) if channels_in_use else 'these channels'}]

Be direct. No bullet points within sections."""
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            with st.spinner('Thinking...'):
                msg = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=500,
                    messages=[{'role': 'user', 'content': prompt}]
                )
            raw = msg.content[0].text
            # Render each section with a styled header
            import re as _re
            sections = _re.split(r'\n(CURRENT ALLOCATION|REBALANCING OPTION|BEST ALLOCATION):\n', raw)
            if len(sections) > 1:
                labels = sections[1::2]
                bodies = sections[2::2]
                colors = {'CURRENT ALLOCATION': '#437CA3', 'REBALANCING OPTION': '#437CA3', 'BEST ALLOCATION': '#1F6152'}
                for label, body in zip(labels, bodies):
                    color = colors.get(label, '#437CA3')
                    st.markdown(f'<div style="background:{color};color:white;padding:6px 10px;border-radius:4px;font-weight:bold;margin-top:8px">{label}</div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="background:#f2f2f2;padding:10px;border-radius:0 0 4px 4px;margin-bottom:4px">{body.strip()}</div>', unsafe_allow_html=True)
            else:
                st.markdown(raw)

with ai_tab_benchmarks:
    st.caption('Plain-English explanations of the benchmark values used in this plan.')

    def _bench_context():
        """Build a reusable benchmark context block for AI prompts."""
        bench_lines = []
        for mkt in (ai_data['selected_markets'] if ai_data else []):
            for goal, chs in (ai_data['goal_channels'].items() if ai_data else []):
                for ch in chs:
                    b = BENCH[mkt][ch]
                    if ch == 'Search':
                        bench_lines.append(f'{MARKET_LABELS[mkt]} / {ch}: CPC €{b["cpc"]:.2f}, CTR {b["ctr"]*100:.1f}%, Click→Session {b.get("click_to_session",0.85)*100:.0f}%')
                    else:
                        bench_lines.append(f'{MARKET_LABELS[mkt]} / {ch}: CPM €{b["cpm"]:.2f}, CTR {b["ctr"]*100:.2f}%, Freq {b.get("frequency",3.0):.1f}' + (f', View Rate {b.get("view_rate",0.31)*100:.0f}%' if ch == 'YouTube' else ''))
        channels_in_use = sorted({ch for chs in (ai_data['goal_channels'].values() if ai_data else []) for ch in chs})
        return bench_lines, channels_in_use

    if st.button('Generate Benchmark Explanations', key='btn_bench'):
        api_key = get_api_key()
        if not api_key:
            st.error('No API key found in .streamlit/secrets.toml')
        else:
            bench_lines, channels_in_use = _bench_context()
            prompt = f"""You are a senior paid media strategist explaining benchmark values to a {audience_type} client in the {industry} sector.

The plan uses these channels: {', '.join(channels_in_use) if channels_in_use else 'various digital channels'}.

Benchmarks in use:
{chr(10).join(bench_lines)}

Write a short explanation (150–180 words) covering:
- What CPM / CPC levels mean in the context of {audience_type} {industry} campaigns — why some markets cost more and whether that premium makes sense for this sector
- Why CTR and View Rate vary between markets and channels — relate this to {audience_type} audience behaviour (e.g. professional vs consumer mindset, intent signals)
- What Frequency means for brand recall in {industry}, and how it should be managed differently depending on whether the goal is awareness or conversion

Write in plain English. No jargon. One paragraph per topic. Frame everything for someone who understands the {industry} business, not a media buyer."""
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            with st.spinner('Thinking...'):
                msg = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=400,
                    messages=[{'role': 'user', 'content': prompt}]
                )
            st.markdown(msg.content[0].text)

    st.divider()
    st.markdown('#### Chat about the benchmarks')
    st.caption('Ask anything — follow up, dig deeper, change direction. The conversation keeps its memory.')

    # Init chat history
    if 'bench_chat' not in st.session_state:
        st.session_state.bench_chat = []

    # Clear button
    if st.session_state.bench_chat:
        if st.button('🗑 Clear conversation', key='bench_clear'):
            st.session_state.bench_chat = []
            st.rerun()

    # Render existing messages
    for msg in st.session_state.bench_chat:
        with st.chat_message(msg['role']):
            st.markdown(msg['content'])

    # Chat input
    if user_input := st.chat_input(
        'e.g. Why is Germany CPM higher than Poland? Can I adjust CTR for a B2B audience?',
        key='bench_chat_input',
    ):
        # Show user message immediately
        with st.chat_message('user'):
            st.markdown(user_input)

        api_key = get_api_key()
        if not api_key:
            with st.chat_message('assistant'):
                st.error('No API key found in .streamlit/secrets.toml')
        else:
            bench_lines, channels_in_use = _bench_context()
            system_ctx = (
                f'You are a senior paid media strategist advising a {audience_type} client '
                f'in the {industry} sector. '
                f'The plan uses these channels: {", ".join(channels_in_use) if channels_in_use else "various digital channels"}. '
                f'Benchmarks in use:\n{chr(10).join(bench_lines)}\n\n'
                f'Answer in 80–150 words. Be direct, reference actual numbers, '
                f'frame for {audience_type} {industry}. No bullet points.'
            )
            # First user turn carries the full context; follow-ups are plain
            api_messages = []
            for i, m in enumerate(st.session_state.bench_chat):
                if i == 0 and m['role'] == 'user':
                    api_messages.append({'role': 'user', 'content': system_ctx + '\n\n---\n\n' + m['content']})
                else:
                    api_messages.append(m)
            # Current turn
            first_turn = len(st.session_state.bench_chat) == 0
            api_messages.append({
                'role': 'user',
                'content': (system_ctx + '\n\n---\n\n' + user_input) if first_turn else user_input,
            })

            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)
            with st.chat_message('assistant'):
                with st.spinner(''):
                    resp = client.messages.create(
                        model='claude-haiku-4-5-20251001',
                        max_tokens=400,
                        messages=api_messages,
                    )
                reply = resp.content[0].text
                st.markdown(reply)

            # Persist both turns
            st.session_state.bench_chat.append({'role': 'user',      'content': user_input})
            st.session_state.bench_chat.append({'role': 'assistant', 'content': reply})
